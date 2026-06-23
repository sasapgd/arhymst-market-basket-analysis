"""Generate supplementary S3-S5 workbooks from reproducible pipeline CSVs."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate S3, S4, and S5 Excel tables.")
    parser.add_argument("--base-dir", type=Path, default=BASE_DIR)
    parser.add_argument(
        "--output-dir", type=Path, default=BASE_DIR / "generated_supplementary_tables"
    )
    return parser.parse_args()


def format_sheet(writer: pd.ExcelWriter, sheet_name: str, title: str) -> None:
    sheet = writer.book[sheet_name]
    sheet.insert_rows(1)
    sheet.cell(1, 1, title)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
    sheet.freeze_panes = "A3"
    for row_number in (1, 2):
        for cell in sheet[row_number]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
    for column_index in range(1, sheet.max_column + 1):
        width = min(
            max(
                len(str(sheet.cell(row, column_index).value or ""))
                for row in range(1, sheet.max_row + 1)
            )
            + 2,
            60,
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def read_csv(path: Path) -> pd.DataFrame:
    if not path.is_file():
        raise FileNotFoundError(path)
    return pd.read_csv(path, sep=";")


def generate_s3(base: Path, output: Path) -> None:
    frame = read_csv(
        base
        / "sensitivity_experiments"
        / "support_confidence"
        / "SUPPORT_CONFIDENCE_SENSITIVITY.csv"
    )[
        [
            "Scenario", "Support", "Confidence", "InitialRules", "RetainedRules",
            "MaxSTNodes", "MaxSTEdges", "SharedEdgesWithBaseline", "TopHub",
        ]
    ]
    frame.columns = [
        "Scenario", "Support", "Confidence", "Initial rules", "Retained rules",
        "MaxST nodes", "MaxST edges", "Shared edges with baseline", "Top hub",
    ]
    path = output / "S3_Table_generated.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="S3 Sensitivity", index=False)
        format_sheet(
            writer,
            "S3 Sensitivity",
            "S3 Table. Sensitivity analysis of support and confidence thresholds.",
        )


def generate_s4(base: Path, output: Path) -> None:
    frame = read_csv(base / "segmented_experiments" / "SEGMENTED_EXPERIMENT_SUMMARY.csv")
    frame["SegmentType"] = frame["SegmentType"].replace(
        {"Gender": "Gender", "Age": "Age", "GenderAge": "Gender × age"}
    )
    frame["Segment"] = frame["Segment"].replace({"F": "Female", "M": "Male"})
    frame["Segment"] = frame["Segment"].str.replace("_", " ", regex=False)
    frame = frame[
        [
            "SegmentType", "Segment", "Baskets", "InitialRules", "RetainedRules",
            "MaxSTNodes", "MaxSTEdges", "SharedEdgesWithBaseline", "TopHub",
        ]
    ]
    frame.columns = [
        "Segment type", "Segment", "Baskets", "Initial rules", "Retained rules",
        "MaxST nodes", "MaxST edges", "Shared edges with baseline", "Top hub",
    ]
    path = output / "S4_Table_generated.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="S4 Segments", index=False)
        format_sheet(
            writer,
            "S4 Segments",
            "S4 Table. Segment-level analysis by gender, age group, and combined gender–age groups.",
        )


def generate_s5(base: Path, output: Path) -> None:
    summary = read_csv(
        base / "statistical_validation" / "statistical_validation_summary.csv"
    )[
        [
            "Maxlen", "RetainedRules", "SignificantAfterBH005",
            "PercentSignificant", "NTransactions",
        ]
    ]
    summary.columns = [
        "Maxlen", "Tested retained rules", "Significant after BH (0.05)",
        "Percent significant", "Transactions",
    ]
    top = read_csv(base / "statistical_validation" / "table5_statistical_validation.csv")
    for column in ("FisherPValue", "BHAdjustedPValue"):
        top[column] = top[column].map(lambda value: "< 0.001" if float(value) < 0.001 else value)

    path = output / "S5_Table_generated.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        top.to_excel(writer, sheet_name="Top_5_rules", index=False)
        format_sheet(
            writer, "Summary", "S5 Table. Statistical validation of retained association rules."
        )
        format_sheet(
            writer,
            "Top_5_rules",
            "S5 Table. Fisher exact and BH-FDR results for the five top-ranked retained rules (maxlen = 3).",
        )


def main() -> None:
    args = parse_args()
    base = args.base_dir.expanduser().resolve()
    output = args.output_dir.expanduser().resolve()
    output.mkdir(parents=True, exist_ok=True)
    generate_s3(base, output)
    generate_s4(base, output)
    generate_s5(base, output)
    print(f"Generated S3-S5 workbooks in: {output}")


if __name__ == "__main__":
    main()
