"""Compare MaxST edge sets and reproduce Supplementary Table S2."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
CRITERIA = ("confidence", "lift", "product")
LABELS = {
    "confidence": "Confidence",
    "lift": "Lift",
    "product": "Lift × Confidence",
}
COMPARISONS = (("confidence", "lift"), ("confidence", "product"), ("lift", "product"))
REQUIRED_COLUMNS = ("Product_1", "Product_2")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare MaxST variants and generate Supplementary Table S2."
    )
    parser.add_argument("--maxlen", nargs="+", type=int, default=[3, 4, 5, 6])
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=BASE_DIR / "timing_runs" / "mst_variants",
        help="Directory containing MST_MAXLEN_<N>_<CRITERION>.csv files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=BASE_DIR / "timing_runs" / "mst_variants",
        help="Directory for S2 CSV and XLSX outputs.",
    )
    return parser.parse_args()


def canonical_edge(product_1: object, product_2: object) -> tuple[str, str]:
    endpoints = sorted((str(product_1).strip(), str(product_2).strip()))
    if not all(endpoints) or endpoints[0] == endpoints[1]:
        raise ValueError(f"Invalid MaxST edge: {product_1!r}, {product_2!r}")
    return endpoints[0], endpoints[1]


def load_edges(path: Path) -> set[tuple[str, str]]:
    if not path.is_file():
        raise FileNotFoundError(f"MaxST file not found: {path}")
    frame = pd.read_csv(path, sep=";")
    missing = [column for column in REQUIRED_COLUMNS if column not in frame]
    if missing:
        raise ValueError(f"{path} is missing columns: {', '.join(missing)}")
    edges = {
        canonical_edge(row.Product_1, row.Product_2)
        for row in frame.itertuples(index=False)
    }
    if len(edges) != len(frame):
        raise ValueError(f"{path} contains duplicate undirected edges.")
    return edges


def _write_csv(frame: pd.DataFrame, path: Path) -> None:
    temporary = path.with_name(f".{path.name}.tmp")
    frame.to_csv(temporary, sep=";", index=False)
    temporary.replace(path)


def _format_sheet(writer: pd.ExcelWriter, sheet_name: str, title: str) -> None:
    worksheet = writer.book[sheet_name]
    worksheet.insert_rows(1)
    worksheet.cell(1, 1, title)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:{worksheet.cell(worksheet.max_row, worksheet.max_column).coordinate}"
    for cell in worksheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    for cell in worksheet[2]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    for column_index in range(1, worksheet.max_column + 1):
        width = min(
            max(
                len(str(worksheet.cell(row_index, column_index).value or ""))
                for row_index in range(1, worksheet.max_row + 1)
            )
            + 2,
            60,
        )
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def main() -> None:
    args = parse_args()
    input_dir = args.input_dir.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    summary_rows: list[dict[str, object]] = []
    detail_rows: list[dict[str, object]] = []

    for maxlen in args.maxlen:
        trees = {
            criterion: load_edges(
                input_dir / f"MST_MAXLEN_{maxlen}_{criterion.upper()}.csv"
            )
            for criterion in CRITERIA
        }
        edge_counts = {len(edges) for edges in trees.values()}
        if len(edge_counts) != 1:
            raise ValueError(f"MaxST variants for maxlen={maxlen} have unequal edge counts.")

        for first, second in COMPARISONS:
            first_only = sorted(trees[first] - trees[second])
            second_only = sorted(trees[second] - trees[first])
            common = trees[first] & trees[second]
            if len(first_only) != len(second_only):
                raise ValueError(
                    f"Unequal substitutions for maxlen={maxlen}, {first} vs {second}."
                )

            comparison = f"{LABELS[first]} vs {LABELS[second]}"
            summary_rows.append(
                {
                    "maxlen": maxlen,
                    "Comparison": comparison,
                    "Common edges": len(common),
                    "Edge substitutions": len(first_only),
                }
            )
            for criterion, edges in ((first, first_only), (second, second_only)):
                for product_1, product_2 in edges:
                    detail_rows.append(
                        {
                            "maxlen": maxlen,
                            "Comparison": comparison,
                            "Present only in": LABELS[criterion],
                            "Product 1": product_1,
                            "Product 2": product_2,
                        }
                    )

    summary = pd.DataFrame(summary_rows)
    details = pd.DataFrame(detail_rows)
    _write_csv(summary, output_dir / "S2_SUMMARY.csv")
    _write_csv(details, output_dir / "S2_DIFFERING_EDGES.csv")

    workbook_path = output_dir / "S2_Table_generated.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        details.to_excel(writer, sheet_name="Differing_edges", index=False)
        _format_sheet(
            writer,
            "Summary",
            "S2 Table. Differing MaxST edges across reduction criteria and maximum rule-length settings.",
        )
        _format_sheet(
            writer,
            "Differing_edges",
            "S2 Table. Detailed list of differing MaxST edges.",
        )

    print(summary.to_string(index=False))
    print(f"Generated workbook: {workbook_path}")


if __name__ == "__main__":
    main()
